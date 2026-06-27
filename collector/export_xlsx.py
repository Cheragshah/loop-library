"""Write the catalog to a formatted Excel workbook: dist/loop-library.xlsx."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DIST = Path(__file__).resolve().parent / "dist"

# (header, loop-key, column width)
COLUMNS = [
    ("Loop name", "name", 34),
    ("Complete loop (command)", "command", 80),
    ("What it does (repurpose for)", "what_it_does", 50),
    ("Category", "category", 16),
    ("Author", "author", 22),
    ("Source", "source", 18),
    ("Trending", "trending", 11),
    ("Votes", "votes", 8),
    ("Mentions", "mentions", 10),
    ("Last seen", "last_seen", 13),
    ("Link", "source_url", 40),
]

HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="F3F4F6")


def export(loops: list[dict]) -> Path:
    DIST.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Trending Loops"
    ws.freeze_panes = "A2"

    for c, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = width

    for r, loop in enumerate(loops, start=2):
        for c, (_, key, _) in enumerate(COLUMNS, start=1):
            val = loop.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(
                vertical="top", wrap_text=key in {"command", "what_it_does", "name"}
            )
            if key == "source_url" and val:
                cell.hyperlink = val
                cell.font = Font(color="2563EB", underline="single")
            if r % 2 == 0:
                if not (key == "source_url" and val):
                    cell.fill = ALT_FILL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(loops) + 1}"
    out = DIST / "loop-library.xlsx"
    wb.save(out)
    return out
